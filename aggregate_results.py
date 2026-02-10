"""
Aggregate all experiment results from logs/FedAvg/MNIST_lenet/ into a single XLSX.
Reads each .txt log, extracts config from header and metrics from the last epoch.
"""
import os
import re
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def parse_header(header_line):
    """Parse the first line (config) of a log file."""
    info = {}
    # Extract key-value pairs
    patterns = {
        'seed': r'seed:\s*(\d+)',
        'epochs': r'epochs:\s*(\d+)',
        'algorithm': r'algorithm:\s*(\w+)',
        'num_clients': r'num_clients:\s*(\d+)',
        'learning_rate': r'learning_rate:\s*([\d.]+)',
        'local_epochs': r'local_epochs:\s*(\d+)',
        'model': r'model:\s*(\w+)',
        'dataset': r'dataset:\s*(\w+)',
        'distribution': r'distribution:\s*([\w-]+)',
        'dirichlet_alpha': r'dirichlet_alpha:\s*([\d.]+)',
        'num_adv': r'num_adv:\s*(\d+)',
        'attack': r"attack:\s*(\w+)",
        'defense': r"defense:\s*(\w+)",
    }
    for key, pattern in patterns.items():
        match = re.search(pattern, header_line)
        if match:
            info[key] = match.group(1)
    return info


def parse_last_epoch(lines):
    """Parse the last Epoch line and extract metrics."""
    metrics = {}
    last_epoch_line = None

    for line in reversed(lines):
        if line.strip().startswith('Epoch'):
            last_epoch_line = line.strip()
            break

    if not last_epoch_line:
        return metrics

    # Extract epoch number
    epoch_match = re.match(r'Epoch\s+(\d+)', last_epoch_line)
    if epoch_match:
        metrics['last_epoch'] = int(epoch_match.group(1))

    # Extract all metric pairs: "Key: Value"
    metric_patterns = {
        'train_acc': r'Train Acc:\s*([\d.]+)',
        'train_loss': r'Train loss:\s*([\d.eE+-]+|nan|inf)',
        'test_acc': r'Test Acc:\s*([\d.]+)',
        'test_loss': r'Test loss:\s*([\d.eE+-]+|nan|inf)',
        'asr': r'ASR:\s*([\d.]+)',
        'asr_loss': r'ASR loss:\s*([\d.eE+-]+|nan|inf)',
    }
    for key, pattern in metric_patterns.items():
        match = re.search(pattern, last_epoch_line)
        if match:
            val = match.group(1)
            if val in ('nan', 'inf'):
                metrics[key] = val
            else:
                metrics[key] = float(val)

    return metrics


def parse_training_time(lines):
    """Extract total training time from the last line."""
    for line in reversed(lines):
        match = re.search(r'using\s+(\d+)\s+minutes?\s+and\s+(\d+)\s+seconds?', line)
        if match:
            minutes = int(match.group(1))
            seconds = int(match.group(2))
            return f"{minutes}m{seconds}s"
    return ''


def parse_status(lines):
    """Determine if training finished or is incomplete."""
    for line in reversed(lines):
        if 'Training finished' in line:
            return 'finished'
    return 'incomplete'


def process_log_file(filepath):
    """Process a single log file and return a dict of results."""
    try:
        with open(filepath, 'r', errors='replace') as f:
            lines = f.readlines()
    except Exception:
        return None

    if not lines:
        return None

    # Parse header (first line)
    header_info = parse_header(lines[0])
    if not header_info.get('attack') or not header_info.get('defense'):
        return None

    # Parse last epoch metrics
    metrics = parse_last_epoch(lines)
    if not metrics:
        return None

    # Parse training time and status
    training_time = parse_training_time(lines)
    status = parse_status(lines)

    # Determine the experiment group from the directory name
    rel_path = os.path.relpath(filepath, LOG_BASE)
    folder = os.path.dirname(rel_path)

    result = {
        'attack': header_info.get('attack', ''),
        'defense': header_info.get('defense', ''),
        'alpha': header_info.get('dirichlet_alpha', ''),
        'num_adv': header_info.get('num_adv', ''),
        'epochs_config': header_info.get('epochs', ''),
        'last_epoch': metrics.get('last_epoch', ''),
        'num_clients': header_info.get('num_clients', ''),
        'lr': header_info.get('learning_rate', ''),
        'test_acc': metrics.get('test_acc', ''),
        'asr': metrics.get('asr', ''),
        'train_acc': metrics.get('train_acc', ''),
        'train_loss': metrics.get('train_loss', ''),
        'test_loss': metrics.get('test_loss', ''),
        'asr_loss': metrics.get('asr_loss', ''),
        'training_time': training_time,
    }
    return result


LOG_BASE = os.path.join(os.path.dirname(__file__), 'logs', 'FedAvg', 'MNIST_lenet')
OUTPUT_XLSX = os.path.join(os.path.dirname(__file__), 'logs', 'FedAvg', 'MNIST_lenet', 'results_summary.xlsx')

COLUMNS = [
    'attack', 'epochs_config', 'alpha', 'num_adv',
    'defense', 'test_acc', 'asr',
    'train_acc', 'train_loss', 'test_loss', 'asr_loss',
    'training_time',
]

# Friendly header names
HEADER_NAMES = [
    'Attack', 'Epochs', 'Alpha', 'Num Adv',
    'Defense', 'Test Acc', 'ASR',
    'Train Acc', 'Train Loss', 'Test Loss', 'ASR Loss',
    'Training Time',
]


def main():
    # Find all .txt log files
    pattern = os.path.join(LOG_BASE, '**', '*.txt')
    txt_files = sorted(glob.glob(pattern, recursive=True))

    results = []
    skipped = []
    for filepath in txt_files:
        result = process_log_file(filepath)
        if result:
            results.append(result)
        else:
            skipped.append(filepath)

    # Sort by: attack -> epochs -> alpha -> num_adv -> defense
    def sort_key(r):
        attack_order = {'DBA': 0, 'BadNets': 1, 'ModelReplacement': 2, 'ALIE': 3, 'IPM': 4}
        return (
            attack_order.get(r['attack'], 99),
            int(r['epochs_config']) if r['epochs_config'] else 0,
            float(r['alpha']) if r['alpha'] else 0,
            int(r['num_adv']) if r['num_adv'] else 0,
            r['defense'],
        )
    results.sort(key=sort_key)

    # Write XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "Results Summary"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    num_font = Font(size=10)

    # Group separator style
    group_font = Font(bold=True, size=11, color="1F4E79")
    group_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    group_align = Alignment(horizontal="left", vertical="center")

    # Alternating row colors for readability
    row_fill_even = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    row_fill_odd = None  # white

    # Highlight columns: green for good Test Acc, red for bad ASR
    # (applied per-cell after writing)

    # Write header
    for col_idx, name in enumerate(HEADER_NAMES, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Group results by scenario (attack + epochs + alpha + num_adv)
    def group_key(r):
        return (r['attack'], r['epochs_config'], r['alpha'], r['num_adv'])

    current_row = 2
    prev_group = None
    data_row_count = 0

    for result in results:
        gk = group_key(result)
        # Insert group separator when scenario changes
        if gk != prev_group:
            if prev_group is not None:
                # Empty separator row
                current_row += 1
            label = f"{gk[0]}  |  {gk[1]} epochs  |  α={gk[2]}  |  {gk[3]} adv"
            cell = ws.cell(row=current_row, column=1, value=label)
            cell.font = group_font
            cell.fill = group_fill
            cell.alignment = group_align
            # Merge across all columns
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=len(COLUMNS))
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=current_row, column=col_idx).fill = group_fill
                ws.cell(row=current_row, column=col_idx).border = thin_border
            current_row += 1
            prev_group = gk
            data_row_count = 0

        # Write data row
        for col_idx, key in enumerate(COLUMNS, 1):
            val = result.get(key, '')
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = num_font
            cell.alignment = center_align
            cell.border = thin_border
            # Alternating row color
            if data_row_count % 2 == 1:
                cell.fill = row_fill_even

        data_row_count += 1
        current_row += 1

    # Auto-adjust column widths
    for col_idx in range(1, len(COLUMNS) + 1):
        max_len = len(str(HEADER_NAMES[col_idx - 1]))
        for row_idx in range(2, current_row):
            cell_val = str(ws.cell(row=row_idx, column=col_idx).value or '')
            max_len = max(max_len, len(cell_val))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 30)

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT_XLSX)
    print(f"Wrote {len(results)} rows to {OUTPUT_XLSX}")
    if skipped:
        print(f"Skipped {len(skipped)} files (no valid data):")
        for s in skipped:
            print(f"  - {os.path.relpath(s, LOG_BASE)}")


if __name__ == '__main__':
    main()
